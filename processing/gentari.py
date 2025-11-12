from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Iterable, Tuple
import requests
import urllib.parse
import pandas as pd
from bs4 import BeautifulSoup
import re
import time
import os
from processing.helper import perform_sentiment_analysis  # add this import at the top

# Source mapping for Type column
SOURCE_MAPPING = {
    'X': 'Social media (X)',
    'Youtube': 'Social media (Youtube)',
    'Facebook Pages': 'Social media (Facebook)',
    'Tiktok': 'Social media (tiktok)',
    'First Party Data': 'Social media (linkedin)',
    'Instagram': 'Social media (instagram)'
}


def clean_text(text):
    """Remove metadata, scripts, extra spaces, and non-content parts."""
    if not text:
        return ""

    # Remove markdown headings and metadata
    text = re.sub(r"^#.*", "", text, flags=re.MULTILINE)
    text = re.sub(r"\[.*?\]\(.*?\)", "", text)  # remove markdown links
    text = re.sub(r"!\[.*?\]\(.*?\)", "", text) # remove markdown images

    # Remove HTML tags and scripts
    soup = BeautifulSoup(text, "html.parser")
    for script in soup(["script", "style", "meta", "noscript", "header", "footer", "nav"]):
        script.decompose()
    text = soup.get_text(separator=" ")

    # Remove extra spaces, line breaks
    text = re.sub(r"\s+", " ", text).strip()
    # Remove UTC timezone format like "(2024-01-01 12:00:00 UTC)"
    text = re.sub(r"\s*\([^)]*UTC[^)]*\)\s*", "", text).strip()

    # Remove non-content boilerplate
    text = re.sub(r"(©|All rights reserved|cookies|subscribe|advertisement|policy).*", "", text, flags=re.I)

    # Keep concise readable text
    return text[:2000]  # limit to 2000 chars for readability


def extract_title_from_content(text):
    """Try to extract a reasonable title from scraped data."""
    if not text:
        return None

    # Markdown heading
    for line in text.splitlines():
        line = line.strip()
        if line.startswith("# "):
            return line.lstrip("# ").strip()

    # HTML title
    try:
        soup = BeautifulSoup(text, "html.parser")
        html_title = soup.title.string if soup.title else None
        if html_title:
            return html_title.strip()
    except Exception:
        pass

    # Fallback: first 8–12 words
    words = text.split()
    return " ".join(words[:12]) if len(words) > 0 else None


# def contains_chinese(text):
#     """Check if text contains Chinese characters."""
#     if not isinstance(text, str):
#         return False
#     return bool(re.search(r'[\u4e00-\u9fff]', text))


# def convert_chinese_to_english(text):
#     """Convert Chinese text to English with pinyin."""
#     if not isinstance(text, str) or text.strip() == "":
#         return text
    
#     if contains_chinese(text):
#         try:
#             from pypinyin import lazy_pinyin, Style
#             from googletrans import Translator
            
#             translator = Translator()
            
#             # Step 1: Convert to Pinyin
#             pinyin = ' '.join(lazy_pinyin(text, style=Style.TONE3))
            
#             # Step 2: Translate Chinese → English
#             translation = translator.translate(text, src='zh-cn', dest='en').text
            
#             # Combine both info
#             return f"{translation} ({pinyin})"
#         except Exception:
#             return text
#     else:
#         return text


# Filtering parameters derived from the reference notebook
UNWANTED_SOURCES = {"blogs", "forums", "reddit"}
ALLOWED_LANGUAGES = {"en", "ms", "zh", "id"}
MEDIA_IMPRESSIONS_THRESHOLD = 5000
SOCIAL_IMPRESSIONS_THRESHOLD = 5000

# Columns required to reproduce the transformation safely
MANDATORY_COLUMNS = {
    "source",
    "language",
    "credibility score",
    "media impressions",
    "social impressions",
    "title",
    "content",
    "date",
    "url",
    "media reach",
    "sentiment class",
}

OUTPUT_COLUMNS = [
    "S.No",
    "Date",
    "Category",
    "Type",
    "Headline",
    "URL",
    "Media Reach",
    "Media Impression",
    "Social Impressions",
    "Sentiment",
]


class ProcessingError(Exception):
    """Raised when Gentari processing fails."""


def read_input_file(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """Load the uploaded Excel or CSV file into a DataFrame."""
    if not file_bytes:
        raise ProcessingError("Uploaded file is empty.")

    suffix = Path(filename).suffix.lower()
    buffer = BytesIO(file_bytes)

    try:
        if suffix in {".xlsx", ".xlsm", ".xls"}:
            return pd.read_excel(buffer)
        if suffix == ".csv":
            return pd.read_csv(buffer)
    except Exception as exc:  # pragma: no cover - depends on pandas internals
        raise ProcessingError(f"Unable to read input file: {exc}") from exc

    raise ProcessingError("Unsupported file type. Please upload an Excel or CSV file.")


def _normalise_columns(df: pd.DataFrame) -> pd.DataFrame:
    normalised = df.copy()
    normalised.columns = [col.strip().lower() for col in normalised.columns]
    return normalised


def _ensure_mandatory_columns(df: pd.DataFrame) -> None:
    missing = MANDATORY_COLUMNS - set(df.columns)
    if missing:
        missing_list = ", ".join(sorted(missing))
        raise ProcessingError(f"Input file is missing required columns: {missing_list}")


def _prepare_numeric(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    return numeric.fillna(0)


def _build_output_filename(original_name: str) -> str:
    base = Path(original_name).stem or "processed"
    return f"{base}_processed.xlsx"


def _format_date(value) -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return "" if value is None else str(value)
    return parsed.strftime("%Y-%m-%d")


def _coalesce_columns(df: pd.DataFrame, candidates: Iterable[str]) -> pd.Series:
    for column in candidates:
        if column in df.columns:
            return df[column].fillna("")
    return pd.Series([""] * len(df), index=df.index)


"""
Categorization logic for Gentari brand need to test it first with the existing system
"""
def categorize_quality_first(row):
    """Categorization logic"""
    topics = str(row.get('topics', '')).lower()
    content = str(row.get('content', '')).lower()
    title = str(row.get('title', '')).lower()
    entities = str(row.get('entities', '')).lower()

    all_text = f"{topics} {content} {title}"
    has_gentari = 'gentari' in all_text
    gentari_prominent = 'gentari' in (title + content)[:300]

    scores = {
        'Gentari RE': 0,
        'Gentari GMOB': 0,
        'Gentari H2': 0,
        'Company News': 0,
        'Industry Mention': 0
    }

    # === H2 ===
    for kw in ['hydrogen',' h2 ','gastech','sarawak','aircraft','aviation',
               'green ammonia','gentari hydrogen','hybrid-electric','fuel cell']:
        if kw in all_text:
            scores['Gentari H2'] += 5

    # === GMOB ===
    for kw in ['ev charging','charging station','ev charger','roaming',
               'idle fee','dc charger','gentari go','shell recharge']:
        if kw in all_text:
            scores['Gentari GMOB'] += 5
    if 'shell' in all_text and 'charging' in all_text:
        scores['Gentari GMOB'] += 6
    if ('electric vehicle' in topics or 'transport' in topics) and 'charging' in all_text:
        scores['Gentari GMOB'] += 4

    # === COMPANY NEWS ===
    for kw in ['pe firms','private equity','stake in gentari','actis',
               'merger','acquisition','valuation']:
        if kw in all_text:
            scores['Company News'] += 8

    has_project = any(p in all_text for p in [
        'solar plant','wind farm','charging station',' mw',' gw','commissioned'
    ])
    if not has_project:
        for kw in [
            'three business pillars','three years of putting clean energy',
            'established in malaysia in september 2022',
            'founded with a clear purpose','marks three years'
        ]:
            if kw in all_text:
                scores['Company News'] += 8
        if 'clean energy arm' in all_text:
            scores['Company News'] += 5
        if ('three years' in all_text or 'tiga tahun' in all_text) and 'gentari' in all_text:
            if 'clean energy' in all_text or 'energy transition' in all_text:
                scores['Company News'] += 5

    # === RE ===
    for kw in ['solar','wind','renewable',' mw',' gw','commissioned',
               'bess','photovoltaic','solar plant','wind farm']:
        if kw in all_text:
            scores['Gentari RE'] += 4
    if 'renewable energy' in topics or 'sustainable energy' in topics:
        scores['Gentari RE'] += 2
    if 'brookfield' in entities:
        scores['Gentari RE'] += 4

    # === INDUSTRY MENTION ===
    for s in ['paling banyak','aplikasi pengecasan','proton e.mas','sebab mengapa',
              'pertamina','dc handal','evcc','competitor','6 aplikasi','market share',
              'best ev','comparison','berbanding']:
        if s in all_text:
            scores['Industry Mention'] += 10
    if has_gentari and not gentari_prominent:
        scores['Industry Mention'] += 5
    companies = ['petronas','shell','actis','pertamina','dc handal']
    if sum(1 for c in companies if c in all_text) >= 2 and not any(kw in all_text for kw in ['pe firms','stake']):
        scores['Industry Mention'] += 5

    max_score = max(scores.values())
    if max_score == 0:
        return 'Industry Mention'
    for cat, score in sorted(scores.items(), key=lambda x: x[1], reverse=True):
        if score == max_score:
            return cat
    return 'Industry Mention'

async def transform_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Transform the raw dataset into the required report structure."""
    work_df = _normalise_columns(df)
    _ensure_mandatory_columns(work_df)

    work_df = work_df.copy()
    work_df["language"] = work_df["language"].astype(str).str.strip().str.lower()
    work_df["source"] = work_df["source"].astype(str).str.strip()
    work_df["title"] = work_df["title"].astype(str)
    work_df["content"] = work_df["content"].astype(str)
    work_df["url"] = work_df["url"].astype(str).str.strip()

    # Fill NaN titles with content (first 100 chars as title)
    title_mask = (work_df["title"].isna()) | (work_df["title"].str.strip() == "") | (work_df["title"] == "nan")
    work_df.loc[title_mask, "title"] = work_df.loc[title_mask, "content"].str[:100]

    # Apply source mapping to Type column
    work_df["source"] = work_df["source"].replace(SOURCE_MAPPING)

    # Scraping logic for missing title/content
    for index, row in work_df.iterrows():
        title_missing = pd.isna(row['title']) or not str(row['title']).strip() or str(row['title']).strip() == "nan"
        content_missing = pd.isna(row['content']) or not str(row['content']).strip() or str(row['content']).strip() == "nan"

        if title_missing and content_missing:
            target_url = str(row['url']).strip()
            if target_url.startswith("http"):
                try:
                    encoded_url = urllib.parse.quote_plus(target_url)
                    BRIGHTDATA_API_KEY = os.getenv("BRIGHTDATA_API_KEY", "")
                    api_url = f"http://api.scrape.do/?token={BRIGHTDATA_API_KEY}&url={encoded_url}&output=markdown"

                    response = requests.get(api_url, timeout=100000)
                    response.raise_for_status()
                    scraped_raw = response.text.strip()

                    if scraped_raw:
                        cleaned_content = clean_text(scraped_raw)
                        scraped_title = extract_title_from_content(scraped_raw)

                        if cleaned_content:
                            work_df.at[index, 'title'] = scraped_title or "Untitled"
                            work_df.at[index, 'content'] = cleaned_content

                    time.sleep(1)  # rate limit: 1 request/second
                except Exception:
                    pass  # Continue if scraping fails

    # Apply Chinese translation to title and content
    # work_df['title'] = work_df['title'].apply(convert_chinese_to_english)
    # work_df['content'] = work_df['content'].apply(convert_chinese_to_english)

    work_df["credibility score"] = _prepare_numeric(work_df["credibility score"])
    work_df["media impressions"] = _prepare_numeric(work_df["media impressions"])
    work_df["social impressions"] = _prepare_numeric(work_df["social impressions"])
    work_df["media reach"] = _prepare_numeric(work_df["media reach"])

    # Stage 1: filter sources, languages, and credibility
    filtered = work_df[
        (~work_df["source"].str.lower().isin(UNWANTED_SOURCES))
        & (work_df["language"].isin(ALLOWED_LANGUAGES))
        & (work_df["credibility score"] == 0)
    ].copy()

    if filtered.empty:
        raise ProcessingError("No rows left after applying the initial filters.")

    # Stage 2: prioritise content with high impressions
    high_media = filtered[filtered["media impressions"] > MEDIA_IMPRESSIONS_THRESHOLD]
    high_social = filtered[filtered["social impressions"] > SOCIAL_IMPRESSIONS_THRESHOLD]
    priority_rows = pd.concat([high_media, high_social]).drop_duplicates()

    remaining = filtered.drop(index=priority_rows.index, errors="ignore").copy()

    priority_rows["_norm_title"] = priority_rows["title"].str.strip().str.lower()
    priority_rows["_norm_content"] = priority_rows["content"].str.strip().str.lower()
    remaining["_norm_title"] = remaining["title"].str.strip().str.lower()
    remaining["_norm_content"] = remaining["content"].str.strip().str.lower()

    title_set = set(priority_rows["_norm_title"].dropna())
    content_set = set(priority_rows["_norm_content"].dropna())

    matched = remaining[
        remaining["_norm_title"].isin(title_set)
        | remaining["_norm_content"].isin(content_set)
    ].copy()

    combined = pd.concat([priority_rows, matched]).drop_duplicates(subset=["url"])
    combined = combined.drop(columns=["_norm_title", "_norm_content"], errors="ignore")
    combined.reset_index(drop=True, inplace=True)

    if combined.empty:
        raise ProcessingError("Processing complete but no rows matched the criteria.")

    # Limit processing to the first 100 rows to keep it short
    # combined = combined.head(20).reset_index(drop=True)
    
    # Ensure the sentiment column exists
    if "sentiment class" not in combined.columns:
        combined["sentiment class"] = ""

    # Iterate rows asynchronously
    for index, row in combined.iterrows():
        try:
            title = str(row.get("title", "")).strip()
            content = str(row.get("content", "")).strip()
            url = str(row.get("url", "")).strip()

            combined_text = f"{title}\n\n{content}".strip()
            if not combined_text:
                continue

            # Run sentiment analysis
            predicted_sentiment = await perform_sentiment_analysis(combined_text, url=url)

            # Clean and compare
            predicted_sentiment = str(predicted_sentiment).strip().lower()
            current_sentiment = str(row.get("sentiment class", "")).strip().lower()

            print(f"Predicted sentiment: {predicted_sentiment} for row :{current_sentiment}") 
            print(f"Current sentiment: {current_sentiment}")
            print(f"Row number: {index + 1}")
            # Update only if the new prediction differs
            if predicted_sentiment and predicted_sentiment != current_sentiment:
                print(f"Updating sentiment for row {index}: {predicted_sentiment}")
                combined.at[index, "sentiment class"] = predicted_sentiment

        except Exception as e:
            print(f"⚠️ Sentiment analysis failed for row {index}: {e}")
            continue


    # Compute category using quality-first categorization
    try:
        category_series = combined.apply(categorize_quality_first, axis=1)
    except Exception:
        category_series = _coalesce_columns(combined, ("topics",))
    headline = combined["title"].replace("", pd.NA).combine_first(combined["content"])

    output = pd.DataFrame(
        {
            "S.No": range(1, len(combined) + 1),
            "Date": combined["date"].apply(_format_date),
            "Category": category_series.astype(str),
            "Type": combined["source"].astype(str),
            "Headline": headline.fillna("").astype(str),
            "URL": combined["url"].astype(str),
            "Media Reach": combined["media reach"],
            "Media Impression": combined["media impressions"],
            "Social Impressions": combined["social impressions"],
            "Sentiment": combined["sentiment class"].astype(str),
        }
    )

    return output[OUTPUT_COLUMNS]


def _sanitize_sheet_name(name: str) -> str:
    invalid = [":", "\\", "/", "?", "*", "[", "]"]
    out = str(name).strip()
    for ch in invalid:
        out = out.replace(ch, " ")
    if not out:
        out = "category"
    return out[:31]


async def process_file(file_bytes: bytes, filename: str) -> Tuple[str, bytes]:
    """Complete pipeline that returns the output filename and Excel bytes."""
    raw_df = read_input_file(file_bytes, filename)
    transformed = await transform_dataframe(raw_df)

    buffer = BytesIO()
    
    # Helper: apply styling to all sheets except excluded ones
    def apply_styling_to_writer(writer, exclude_sheets=None):
        if exclude_sheets is None:
            exclude_sheets = set()

        try:
            from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
        except Exception:
            # openpyxl not available or import failed — skip styling
            return

        header_fill = PatternFill(fill_type='solid', start_color='FFADD8E6')
        header_font = Font(bold=True, color='FF1F4E78')
        light_blue_fill = PatternFill(fill_type='solid', start_color='FFF0F4F8')
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for sheet_name, ws in writer.sheets.items():
            if sheet_name in exclude_sheets:
                continue
            try:
                max_col = ws.max_column
                max_row = ws.max_row

                # default widths for first 9 columns, fallback for extras
                default_widths = {
                    1: 10.0,   # A
                    2: 30.0,  # B
                    3: 30.0,  # C
                    4: 30.0,  # D
                    5: 45.0,  # E
                    6: 45.0,  # F
                    7: 30.0,  # G
                    8: 30.0,  # H
                    9: 30.0   # I
                }

                for col_idx in range(1, max_col + 1):
                    col_letter = get_column_letter(col_idx)
                    width = default_widths.get(col_idx, 20.0)
                    ws.column_dimensions[col_letter].width = width

                # Apply borders and styling to all cells
                for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                # Header styling — first row
                header_row = 1
                for cell in ws[header_row]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Apply light blue fill to column D if present
                if max_col >= 4:
                    for cell in ws['D']:
                        if cell.row > 1:  # Skip header row
                            cell.fill = light_blue_fill

            except Exception:
                import logging
                logging.getLogger(__name__).exception(f"Failed to style sheet: {sheet_name}")
                continue

    # Build multi-sheet workbook: data, summary, data_chart (with charts)
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Main filtered data
        transformed.to_excel(writer, index=False, sheet_name="Complete sheet")

        # Compute basic aggregates
        sentiment_counts = (
            transformed["Sentiment"].value_counts().rename_axis("Sentiment").reset_index(name="Count")
        )
        type_counts = (
            transformed["Type"].value_counts().rename_axis("Type").reset_index(name="Count")
        )
        totals_df = pd.DataFrame(
            {
                "Metric": [
                    "Total Rows",
                    "Total Media Impression",
                    "Total Media Reach",
                ],
                "Value": [
                    len(transformed),
                    float(transformed["Media Impression"].sum()),
                    float(transformed["Media Reach"].sum()),
                ],
            }
        )

        # Simple insights
        top_type = type_counts.iloc[0]["Type"] if not type_counts.empty else "-"
        top_sentiment = (
            sentiment_counts.iloc[0]["Sentiment"] if not sentiment_counts.empty else "-"
        )
        insights = pd.DataFrame(
            {
                "Insights": [
                    f"Most frequent type: {top_type}",
                    f"Dominant sentiment: {top_sentiment}",
                    "Focus on high-reach sources to amplify visibility.",
                    "Monitor negative sentiment items for mitigation.",
                ]
            }
        )

        # Write summary sheet (tables stacked)
        start = 0
        sentiment_counts.to_excel(writer, index=False, sheet_name="summary", startrow=start)
        start += len(sentiment_counts) + 3
        type_counts.to_excel(writer, index=False, sheet_name="summary", startrow=start)
        start += len(type_counts) + 3
        totals_df.to_excel(writer, index=False, sheet_name="summary", startrow=start)
        start += len(totals_df) + 3
        insights.to_excel(writer, index=False, sheet_name="summary", startrow=start)

        # Write data for charts
        sentiment_counts.to_excel(writer, index=False, sheet_name="data_chart", startrow=0, startcol=0)
        type_counts.to_excel(writer, index=False, sheet_name="data_chart", startrow=0, startcol=4)

        # Add charts using openpyxl
        wb = writer.book
        ws = writer.sheets["data_chart"]

        # Additional matplotlib charts with annotations embedded as images
        try:
            import matplotlib.pyplot as plt
            import tempfile
            from openpyxl.drawing.image import Image as XLImage

            temp_paths = []

            # Annotated bar chart for Type counts (improved layout)
            if not type_counts.empty:
                n_types = len(type_counts)
                fig_w = max(6, min(14, 0.7 * n_types))
                fig_h = 4.0
                fig, ax = plt.subplots(figsize=(fig_w, fig_h), constrained_layout=True)

                bars = ax.bar(type_counts['Type'], type_counts['Count'], color="#2E5B8A")
                ax.set_title('Type Count', fontsize=12)
                ax.set_ylabel('Count', fontsize=10)
                ax.set_xlabel('Type', fontsize=10)

                # Smaller fonts and rotation to prevent overlap
                if n_types > 6:
                    rotation, ha, fontsize = 45, 'right', 8
                else:
                    rotation, ha, fontsize = 0, 'center', 9
                ax.set_xticklabels(type_counts['Type'], rotation=rotation, ha=ha, fontsize=fontsize)
                ax.tick_params(axis='y', labelsize=9)

                # Annotate bars with smaller font
                max_val = type_counts['Count'].max() if not type_counts['Count'].empty else 0
                offset = max(2, 0.03 * max_val)
                for rect in bars:
                    height = rect.get_height()
                    ax.annotate(
                        f'{int(height):,}',
                        xy=(rect.get_x() + rect.get_width() / 2, height),
                        xytext=(0, offset),
                        textcoords="offset points",
                        ha='center', va='bottom', fontsize=8
                    )

                # tighten layout and save
                tmp1 = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                fig.savefig(tmp1.name, dpi=150, bbox_inches='tight')
                plt.close(fig)
                temp_paths.append(tmp1.name)
                ws.add_image(XLImage(tmp1.name), 'A22')

            # Annotated pie chart for Sentiment (improved label handling)
            if not sentiment_counts.empty:
                n_slices = len(sentiment_counts)
                fig_w = 6
                fig_h = 4.5
                fig, ax = plt.subplots(figsize=(fig_w, fig_h), constrained_layout=True)

                counts = sentiment_counts['Count'].tolist()
                labels = sentiment_counts['Sentiment'].tolist()
                total = sum(counts)

                # Create pie without labels to avoid overlap
                wedges, texts, autotexts = ax.pie(
                    counts,
                    labels=None,
                    autopct='',
                    startangle=90
                )
                ax.set_title('Sentiment Share', fontsize=12)
                
                # Create legend with count and percentage
                legend_labels = [f"{label}: {count} ({count/total*100:.1f}%)" for label, count in zip(labels, counts)]
                ax.legend(wedges, legend_labels, title="Sentiment", bbox_to_anchor=(1.15, 0.5), loc='center left', fontsize=8, title_fontsize=9)

                tmp2 = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                fig.savefig(tmp2.name, dpi=150, bbox_inches='tight')
                plt.close(fig)
                temp_paths.append(tmp2.name)
                ws.add_image(XLImage(tmp2.name), 'E22')
        except Exception:
            pass


        # One sheet per Category with same columns as transformed
        existing_names = set(writer.sheets.keys())
        for cat in sorted(transformed["Category"].dropna().unique()):
            base = _sanitize_sheet_name(f"{cat}")
            name = base
            suffix = 1
            while name in existing_names:
                name = _sanitize_sheet_name(f"{base}_{suffix}")
                suffix += 1
            
            # Get category data and reset S.No for this sheet
            cat_data = transformed[transformed["Category"] == cat].copy()
            cat_data["S.No"] = range(1, len(cat_data) + 1)
            cat_data.to_excel(writer, index=False, sheet_name=name)
            existing_names.add(name)

        # Apply styling to all sheets (including summary and data_chart)
        apply_styling_to_writer(writer, exclude_sheets=set())

    buffer.seek(0)
    return _build_output_filename(filename), buffer.getvalue()


